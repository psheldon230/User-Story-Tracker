import io
import csv
import json
import logging
from datetime import date, datetime

import openpyxl
from openpyxl.styles import PatternFill
from flask import Flask, jsonify, render_template, request, send_file, redirect, url_for

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


# ── Jira CSV parser ───────────────────────────────────────────────────────────

KEY_CANDIDATES    = ["Issue key", "Issue Key", "Key", "key"]
SPRINT_CANDIDATES = ["Sprint", "sprint", "Sprint Name", "Custom field (Sprint)"]


def _normalise_sprint(raw: str) -> str:
    """Convert Jira sprint codes like 'Y26.SP04' or 'SP04' to 'Sprint 4'."""
    import re
    if not raw:
        return ""
    m = re.search(r"SP0*(\d+)", raw, re.IGNORECASE)
    if m:
        return f"Sprint {int(m.group(1))}"
    m2 = re.search(r"sprint\s*(\d+)", raw, re.IGNORECASE)
    if m2:
        return f"Sprint {int(m2.group(1))}"
    return raw


def _find_column(header_row: list, candidates: list):
    lower = [h.strip().lower() for h in header_row]
    for candidate in candidates:
        idx = next((i for i, h in enumerate(lower) if h == candidate.lower()), None)
        if idx is not None:
            return idx
    return None


def _find_sprint_indices(header_row: list, candidates: list):
    """Return ALL column indices that match a sprint candidate."""
    lower = [h.strip().lower() for h in header_row]
    indices = []
    for candidate in candidates:
        for i, h in enumerate(lower):
            if h == candidate.lower():
                indices.append(i)
    return indices


def _best_sprint_for_row(row: list, sprint_indices: list) -> str:
    """Among all sprint columns, return the normalised value with the highest sprint number."""
    import re

    def extract_num(val):
        if not val:
            return -1
        m = re.search(r'SP0*(\d+)', val, re.IGNORECASE)
        if m:
            return int(m.group(1))
        m2 = re.search(r'sprint\s*(\d+)', val, re.IGNORECASE)
        if m2:
            return int(m2.group(1))
        return -1

    best_val, best_num = "", -1
    for idx in sprint_indices:
        val = row[idx].strip() if idx < len(row) else ""
        n = extract_num(val)
        if n > best_num:
            best_num, best_val = n, val
    return _normalise_sprint(best_val)


def parse_jira_csv(csv_bytes: bytes) -> tuple:
    text = csv_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)

    if len(rows) < 2:
        raise RuntimeError("Jira CSV appears to be empty or has no data rows.")

    header = rows[0]
    key_idx       = _find_column(header, KEY_CANDIDATES)
    sprint_indices = _find_sprint_indices(header, SPRINT_CANDIDATES)

    warnings = []
    if key_idx is None:
        raise RuntimeError(
            f"Could not find an Issue Key column in the CSV. "
            f"Columns found: {', '.join(header[:10])}"
        )
    if not sprint_indices:
        warnings.append("No Sprint column found — Sprint will be left blank.")

    import re
    _jira_key_re = re.compile(r"^[A-Z][A-Z0-9_]+-\d+$", re.IGNORECASE)

    stories = []
    for row in rows[1:]:
        if not row:
            continue
        key = row[key_idx].strip() if key_idx < len(row) else ""
        if not key or not _jira_key_re.match(key):
            continue
        sprint = _best_sprint_for_row(row, sprint_indices) if sprint_indices else ""
        stories.append({"key": key, "sprint": sprint})

    return stories, warnings


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _is_today_header(value) -> bool:
    today = date.today()
    if isinstance(value, datetime):
        return value.date() == today
    if isinstance(value, date):
        return value == today
    if isinstance(value, str):
        for fmt in (
            "%m/%d/%Y", "%-m/%-d/%Y", "%Y-%m-%d",
            "%m-%d-%Y", "%d/%m/%Y", "%B %d, %Y", "%b %d, %Y",
        ):
            try:
                if datetime.strptime(value.strip(), fmt).date() == today:
                    return True
            except ValueError:
                continue
    return False


def _is_test_user_header(value) -> bool:
    if not value:
        return False
    return str(value).strip().lower() in (
        "test user", "testuser", "tester", "test_user"
    )


_BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")


def _write_story_row(ws, row_idx, story, max_col, today_col, test_user_col, today):
    ws.cell(row=row_idx, column=1).value = story["sprint"]
    ws.cell(row=row_idx, column=2).value = story["key"]
    if today_col:
        ws.cell(row=row_idx, column=today_col).value = today
    if test_user_col:
        ws.cell(row=row_idx, column=test_user_col).value = "peter"


def sync_excel(excel_bytes: bytes, stories: list) -> tuple:
    from collections import defaultdict

    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    sheet_name = "Functional Testing"
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
        )
    ws = wb[sheet_name]

    # Read existing keys and the last row index per sprint
    existing_keys: set = set()
    sprint_last_row: dict = {}  # sprint value -> last row index (1-based)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if len(row) > 1 and row[1] is not None:
            existing_keys.add(str(row[1]).strip())
        sprint_val = str(row[0]).strip() if row[0] else ""
        if sprint_val:
            sprint_last_row[sprint_val] = row_idx

    today_col = None
    test_user_col = None
    for cell in ws[1]:
        if cell.value is None:
            continue
        if today_col is None and _is_today_header(cell.value):
            today_col = cell.column
        if test_user_col is None and _is_test_user_header(cell.value):
            test_user_col = cell.column

    max_col = ws.max_column
    today = date.today()
    added_keys = []
    skipped_keys = []

    # Partition into old-sprint (insert near existing) and new-sprint (append with separator)
    old_sprint_stories = defaultdict(list)
    new_sprint_stories = []

    for story in stories:
        if story["key"] in existing_keys:
            skipped_keys.append(story["key"])
            continue
        existing_keys.add(story["key"])  # prevent within-batch duplicates
        if story["sprint"] in sprint_last_row:
            old_sprint_stories[story["sprint"]].append(story)
        else:
            new_sprint_stories.append(story)

    # Insert old-sprint stories after the last row of their sprint.
    # Process bottom-to-top so earlier insertions don't shift later targets.
    for sprint in sorted(old_sprint_stories, key=lambda s: sprint_last_row[s], reverse=True):
        group = old_sprint_stories[sprint]
        insert_after = sprint_last_row[sprint]
        ws.insert_rows(insert_after + 1, amount=len(group))
        for i, story in enumerate(group):
            _write_story_row(ws, insert_after + 1 + i, story, max_col, today_col, test_user_col, today)
            added_keys.append(story["key"])

    # Append new-sprint stories with a black separator row first
    if new_sprint_stories:
        ws.append([None] * max_col)
        sep_row = ws.max_row
        for col in range(1, max_col + 1):
            ws.cell(row=sep_row, column=col).fill = _BLACK_FILL

        for story in new_sprint_stories:
            new_row = [None] * max_col
            new_row[0] = story["sprint"]
            new_row[1] = story["key"]
            if today_col is not None:
                new_row[today_col - 1] = today
            if test_user_col is not None:
                new_row[test_user_col - 1] = "peter"
            ws.append(new_row)
            added_keys.append(story["key"])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read(), added_keys, skipped_keys


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    error = request.args.get("error", "")
    return render_template("index.html", error=error)


@app.route("/parse-csv", methods=["POST"])
def parse_csv():
    """VDI step: parse Jira CSV and return a small JSON file of stories."""
    try:
        csv_file = request.files.get("jira_csv")
        if not csv_file or csv_file.filename == "":
            return redirect(url_for("index", error="Please upload your Jira CSV export."), 303)

        stories, warnings = parse_jira_csv(csv_file.read())
        if not stories:
            return redirect(url_for("index", error="No stories found in the Jira CSV."), 303)

        today_str = date.today().strftime("%Y%m%d")
        json_bytes = json.dumps(stories, indent=2).encode("utf-8")
        return send_file(
            io.BytesIO(json_bytes),
            mimetype="application/json",
            as_attachment=True,
            download_name=f"stories_{today_str}.json",
        )

    except RuntimeError as e:
        logger.warning("Handled error: %s", e)
        return redirect(url_for("index", error=str(e)), 303)
    except Exception as e:
        logger.exception("Unexpected error")
        return redirect(url_for("index", error=f"Unexpected error: {e}"), 303)


@app.route("/sync", methods=["POST"])
def sync():
    """Sync stories into Excel. Accepts either a Jira CSV or a pre-parsed JSON file."""
    try:
        csv_file   = request.files.get("jira_csv")
        json_file  = request.files.get("stories_json")
        excel_file = request.files.get("excel_file")

        if not excel_file or excel_file.filename == "":
            return redirect(url_for("index", error="Please upload your Excel tracker file."), 303)

        json_text  = request.form.get("stories_json_text", "").strip()

        warnings = []
        if json_file and json_file.filename != "":
            try:
                stories = json.loads(json_file.read().decode("utf-8"))
            except Exception:
                return redirect(url_for("index", error="Could not parse the stories JSON file."), 303)
            if not stories:
                return redirect(url_for("index", error="No stories found in the JSON file."), 303)
        elif json_text:
            try:
                stories = json.loads(json_text)
            except Exception:
                return redirect(url_for("index", error="Could not parse the pasted JSON. Make sure you copied the full text."), 303)
            if not stories:
                return redirect(url_for("index", error="No stories found in the pasted JSON."), 303)
        elif csv_file and csv_file.filename != "":
            stories, warnings = parse_jira_csv(csv_file.read())
            if not stories:
                return redirect(url_for("index", error="No stories found in the Jira CSV."), 303)
        else:
            return redirect(url_for("index", error="Please upload a Jira CSV, a stories JSON file, or paste the JSON text."), 303)

        import base64
        updated_bytes, added_keys, skipped_keys = sync_excel(excel_file.read(), stories)

        today_str = date.today().strftime("%Y%m%d")
        filename = f"stories_updated_{today_str}.xlsx"

        # AJAX callers get JSON summary + base64 Excel
        if request.headers.get("X-Requested-With") == "fetch":
            return jsonify({
                "added": added_keys,
                "skipped": skipped_keys,
                "filename": filename,
                "excel_b64": base64.b64encode(updated_bytes).decode("utf-8"),
            })

        return send_file(
            io.BytesIO(updated_bytes),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except RuntimeError as e:
        logger.warning("Handled error: %s", e)
        return redirect(url_for("index", error=str(e)), 303)
    except Exception as e:
        logger.exception("Unexpected error")
        return redirect(url_for("index", error=f"Unexpected error: {e}"), 303)


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
