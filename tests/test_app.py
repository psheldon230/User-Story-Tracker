"""Tests for app.py — CSV parser, Excel sync, and Flask routes."""
import io
import json
from datetime import date, datetime, timedelta

import openpyxl
import pytest

from openpyxl.styles import PatternFill

from app import app, parse_jira_csv, sync_excel


# ── Helpers ───────────────────────────────────────────────────────────────────

def csv_bytes(*lines):
    return "\n".join(lines).encode("utf-8")


def make_excel(headers, rows=None):
    """Return bytes of a minimal .xlsx with a 'Functional Testing' sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Functional Testing"
    ws.append(headers)
    for row in (rows or []):
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def make_excel_with_trailing_blanks(headers, rows, extra_blank_rows=10):
    """Simulates an Excel table that has styled empty rows below the data,
    which is what happens with real Excel table objects (ListObjects)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Functional Testing"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    last_data = ws.max_row
    white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for i in range(1, extra_blank_rows + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=last_data + i, column=col).fill = white
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── parse_jira_csv ────────────────────────────────────────────────────────────

class TestParseJiraCsv:

    def test_basic_six_stories(self):
        data = csv_bytes(
            "Issue key,Summary,Sprint",
            "PROJ-1,Story one,Sprint 1",
            "PROJ-2,Story two,Sprint 1",
            "PROJ-3,Story three,Sprint 2",
            "PROJ-4,Story four,Sprint 2",
            "PROJ-5,Story five,Sprint 3",
            "PROJ-6,Story six,Sprint 3",
        )
        stories, warnings = parse_jira_csv(data)
        assert len(stories) == 6
        assert stories[0] == {"key": "PROJ-1", "sprint": "Sprint 1"}
        assert stories[5] == {"key": "PROJ-6", "sprint": "Sprint 3"}

    def test_filters_out_non_key_rows(self):
        """Rows like 'Product Owner', metadata, blank keys must be dropped."""
        data = csv_bytes(
            "Issue key,Summary,Sprint",
            "PROJ-1,Real story,Sprint 1",
            "Product Owner,Some metadata,",
            ",Empty key row,Sprint 1",
            "not-a-key,bad format,Sprint 1",
            "PROJ-2,Another real story,Sprint 2",
            "123,pure number,Sprint 1",
            "PROJ-3,Third story,Sprint 1",
        )
        stories, _ = parse_jira_csv(data)
        keys = [s["key"] for s in stories]
        assert keys == ["PROJ-1", "PROJ-2", "PROJ-3"]

    def test_empty_rows_skipped(self):
        data = csv_bytes(
            "Issue key,Sprint",
            "PROJ-10,Sprint 1",
            "",
            "PROJ-11,Sprint 2",
        )
        stories, _ = parse_jira_csv(data)
        assert len(stories) == 2

    def test_no_sprint_column_gives_warning(self):
        data = csv_bytes(
            "Issue key,Summary",
            "PROJ-1,Story one",
        )
        stories, warnings = parse_jira_csv(data)
        assert len(stories) == 1
        assert stories[0]["sprint"] == ""
        assert any("sprint" in w.lower() for w in warnings)

    def test_missing_key_column_raises(self):
        data = csv_bytes(
            "Summary,Status",
            "Some story,Open",
        )
        with pytest.raises(RuntimeError, match="Issue Key"):
            parse_jira_csv(data)

    def test_empty_csv_raises(self):
        with pytest.raises(RuntimeError, match="empty"):
            parse_jira_csv(b"Issue key\n")  # header only, no data

    def test_totally_empty_raises(self):
        with pytest.raises(RuntimeError):
            parse_jira_csv(b"")

    def test_alternate_header_names(self):
        """Accepts 'Key' and 'Custom field (Sprint)' variants."""
        data = csv_bytes(
            "Key,Summary,Custom field (Sprint)",
            "ABC-42,Story,Sprint 5",
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["key"] == "ABC-42"
        assert stories[0]["sprint"] == "Sprint 5"

    def test_case_insensitive_key_matching(self):
        data = csv_bytes(
            "issue key,Sprint",
            "PROJ-99,Sprint 1",
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["key"] == "PROJ-99"

    def test_utf8_bom_handled(self):
        raw = "Issue key,Sprint\nPROJ-1,Sprint 1\n"
        stories, _ = parse_jira_csv(raw.encode("utf-8-sig"))  # encode adds the BOM
        assert stories[0]["key"] == "PROJ-1"

    def test_key_with_underscores_and_numbers(self):
        """Project codes like 'MY_PROJ2-10' should be valid."""
        data = csv_bytes(
            "Issue key,Sprint",
            "MY_PROJ2-10,Sprint 1",
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["key"] == "MY_PROJ2-10"

    def test_sprint_normalised(self):
        """Sprint codes like 'Y26.SP04' are normalised to 'Sprint 4'."""
        data = csv_bytes(
            "Issue key,Sprint",
            "PROJ-1,Y26.SP04",
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Sprint 4"


# ── sync_excel ────────────────────────────────────────────────────────────────

class TestSyncExcel:

    def _read_excel_rows(self, excel_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active
        return list(ws.iter_rows(values_only=True))

    def test_new_stories_appended(self):
        today_str = date.today().strftime("%-m/%-d/%Y")
        xl = make_excel(["Sprint", "Key", today_str])
        stories = [{"key": "PROJ-1", "sprint": "Sprint 1"},
                   {"key": "PROJ-2", "sprint": "Sprint 2"}]
        result, added, skipped = sync_excel(xl, stories)
        assert added == ["PROJ-1", "PROJ-2"]
        assert skipped == []
        rows = self._read_excel_rows(result)
        keys = [r[1] for r in rows[1:]]
        assert "PROJ-1" in keys
        assert "PROJ-2" in keys

    def test_existing_keys_skipped(self):
        today_str = date.today().strftime("%-m/%-d/%Y")
        xl = make_excel(["Sprint", "Key", today_str],
                        rows=[["Sprint 1", "PROJ-1", None]])
        stories = [{"key": "PROJ-1", "sprint": "Sprint 1"},
                   {"key": "PROJ-2", "sprint": "Sprint 2"}]
        _, added, skipped = sync_excel(xl, stories)
        assert added == ["PROJ-2"]
        assert skipped == ["PROJ-1"]

    def test_no_duplicates_within_batch(self):
        xl = make_excel(["Sprint", "Key"])
        stories = [{"key": "PROJ-1", "sprint": "S1"},
                   {"key": "PROJ-1", "sprint": "S1"}]  # duplicate
        _, added, skipped = sync_excel(xl, stories)
        assert added == ["PROJ-1"]
        assert skipped == ["PROJ-1"]

    def test_sprint_written_to_column_a(self):
        """New sprint gets a black separator row then the story."""
        xl = make_excel(["Sprint", "Key"])
        stories = [{"key": "PROJ-7", "sprint": "Sprint 3"}]
        result, _, _ = sync_excel(xl, stories)
        rows = self._read_excel_rows(result)
        # rows[1] is the black separator, rows[2] is the story
        assert rows[2][0] == "Sprint 3"
        assert rows[2][1] == "PROJ-7"

    def test_old_sprint_inserted_near_existing(self):
        """Stories from a sprint already in the sheet are inserted after existing sprint rows."""
        xl = make_excel(["Sprint", "Key"], rows=[
            ["Sprint 4", "PROJ-1"],
            ["Sprint 4", "PROJ-2"],
        ])
        stories = [{"key": "PROJ-3", "sprint": "Sprint 4"}]
        result, added, _ = sync_excel(xl, stories)
        assert added == ["PROJ-3"]
        rows = self._read_excel_rows(result)
        # Should be header, PROJ-1, PROJ-2, PROJ-3 — no black row
        assert rows[3][1] == "PROJ-3"
        assert len(rows) == 4  # no black separator added

    def test_new_sprint_gets_black_separator(self):
        """A brand-new sprint gets a black separator row before it."""
        xl = make_excel(["Sprint", "Key"], rows=[["Sprint 4", "PROJ-1"]])
        stories = [{"key": "PROJ-10", "sprint": "Sprint 5"}]
        result, added, _ = sync_excel(xl, stories)
        assert added == ["PROJ-10"]
        rows = self._read_excel_rows(result)
        # header, PROJ-1, black row, PROJ-10
        assert len(rows) == 4
        # black row: all cells None (values_only strips fill, so just check story position)
        assert rows[3][1] == "PROJ-10"

    def test_mixed_old_and_new_sprint(self):
        """Old-sprint stories inserted inline; new-sprint stories get black separator."""
        xl = make_excel(["Sprint", "Key"], rows=[
            ["Sprint 4", "PROJ-1"],
        ])
        stories = [
            {"key": "PROJ-2", "sprint": "Sprint 4"},   # old sprint
            {"key": "PROJ-10", "sprint": "Sprint 5"},  # new sprint
        ]
        result, added, _ = sync_excel(xl, stories)
        assert set(added) == {"PROJ-2", "PROJ-10"}
        rows = self._read_excel_rows(result)
        # header, PROJ-1, PROJ-2 (inserted), black row, PROJ-10
        assert len(rows) == 5
        assert rows[2][1] == "PROJ-2"
        assert rows[4][1] == "PROJ-10"

    def test_empty_stories_list(self):
        xl = make_excel(["Sprint", "Key"])
        _, added, skipped = sync_excel(xl, [])
        assert added == []
        assert skipped == []


# ── Flask routes ──────────────────────────────────────────────────────────────

@pytest.fixture
def client():
    app.config["TESTING"] = True
    with app.test_client() as c:
        yield c


class TestParseCsvRoute:

    def test_returns_json_file(self, client):
        data = csv_bytes(
            "Issue key,Sprint",
            "PROJ-1,Sprint 1",
            "PROJ-2,Sprint 2",
        )
        resp = client.post(
            "/parse-csv",
            data={"jira_csv": (io.BytesIO(data), "export.csv")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        assert resp.content_type == "application/json"
        stories = json.loads(resp.data)
        assert len(stories) == 2
        assert stories[0]["key"] == "PROJ-1"

    def test_no_file_redirects_with_error(self, client):
        resp = client.post("/parse-csv", data={}, content_type="multipart/form-data")
        assert resp.status_code == 303
        assert "error" in resp.headers["Location"]

    def test_dummy_rows_excluded_from_json(self, client):
        data = csv_bytes(
            "Issue key,Sprint",
            "PROJ-1,Sprint 1",
            "Product Owner,,",
            "PROJ-2,Sprint 2",
        )
        resp = client.post(
            "/parse-csv",
            data={"jira_csv": (io.BytesIO(data), "export.csv")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        stories = json.loads(resp.data)
        assert len(stories) == 2


class TestSyncRoute:

    def _make_multipart(self, stories, xl_bytes):
        return {
            "stories_json": (io.BytesIO(json.dumps(stories).encode()), "stories.json"),
            "excel_file": (io.BytesIO(xl_bytes), "tracker.xlsx"),
        }

    def test_sync_via_json(self, client):
        xl = make_excel(["Sprint", "Key"])
        stories = [{"key": "PROJ-1", "sprint": "Sprint 1"}]
        resp = client.post(
            "/sync",
            data=self._make_multipart(stories, xl),
            content_type="multipart/form-data",
        )
        assert resp.status_code == 200
        assert "spreadsheet" in resp.content_type

    def test_missing_excel_redirects(self, client):
        resp = client.post(
            "/sync",
            data={"stories_json": (io.BytesIO(b"[]"), "s.json")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 303

    def test_no_input_redirects(self, client):
        xl = make_excel(["Sprint", "Key"])
        resp = client.post(
            "/sync",
            data={"excel_file": (io.BytesIO(xl), "tracker.xlsx")},
            content_type="multipart/form-data",
        )
        assert resp.status_code == 303

    def test_fetch_returns_summary_json(self, client):
        xl = make_excel(["Sprint", "Key"],
                        rows=[["Sprint 1", "PROJ-1", None]])
        stories = [
            {"key": "PROJ-1", "sprint": "Sprint 1"},  # already exists
            {"key": "PROJ-2", "sprint": "Sprint 2"},  # new
        ]
        resp = client.post(
            "/sync",
            data=self._make_multipart(stories, xl),
            content_type="multipart/form-data",
            headers={"X-Requested-With": "fetch"},
        )
        assert resp.status_code == 200
        assert resp.content_type == "application/json"
        data = json.loads(resp.data)
        assert data["added"] == ["PROJ-2"]
        assert data["skipped"] == ["PROJ-1"]
        assert "excel_b64" in data
        assert "filename" in data


class TestDuplicateSprintColumn:

    def test_picks_higher_sprint_per_row(self):
        """Each row independently gets the higher of its two sprint values."""
        data = csv_bytes(
            "Issue key,Sprint,Summary,Sprint",
            "PROJ-1,Sprint 4,Story one,Sprint 5",   # row has 4 and 5 → must pick 5
            "PROJ-2,Sprint 5,Story two,Sprint 4",   # row has 5 and 4 → must pick 5
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Sprint 5"
        assert stories[1]["sprint"] == "Sprint 5"

    def test_different_sprints_per_row(self):
        """Rows with different sprint values each get their own max."""
        data = csv_bytes(
            "Issue key,Sprint,Summary,Sprint",
            "PROJ-1,Sprint 3,Story one,Sprint 6",
            "PROJ-2,Sprint 7,Story two,Sprint 2",
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Sprint 6"
        assert stories[1]["sprint"] == "Sprint 7"

    def test_jira_sprint_code_format_not_confused_by_year(self):
        """Y26.SP04 should resolve to Sprint 4, not Sprint 26."""
        data = csv_bytes(
            "Issue key,Sprint,Summary,Sprint",
            "PROJ-1,Y26.SP03,Story one,Y26.SP05",
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Sprint 5"

    def test_mixed_empty_sprint_columns(self):
        """If one sprint column is empty and one has values, use the one with values."""
        data = csv_bytes(
            "Issue key,Sprint,Summary,Sprint",
            "PROJ-1,,Story one,Sprint 4",
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Sprint 4"

    def test_many_columns_before_sprint(self):
        """Sprint columns deep in the CSV (like column AB/AC) are still found."""
        cols = ["col" + str(i) for i in range(26)]
        header = "Issue key," + ",".join(cols) + ",Sprint,Extra,Sprint"
        row    = "PROJ-1,"    + ",".join(["x"] * 26) + ",Sprint 3,y,Sprint 7"
        data = csv_bytes(header, row)
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Sprint 7"

    def test_quoted_fields_with_commas_dont_shift_columns(self):
        data = csv_bytes(
            'Issue key,Sprint,Summary,Sprint',
            '"PROJ-1",Sprint 2,"Summary with, a comma",Sprint 6',
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["key"] == "PROJ-1"
        assert stories[0]["sprint"] == "Sprint 6"

    def test_quoted_fields_with_newlines_dont_shift_columns(self):
        raw = b'Issue key,Sprint,Summary,Sprint\r\n"PROJ-1",Sprint 2,"Line one\nLine two",Sprint 8\r\n'
        stories, _ = parse_jira_csv(raw)
        assert stories[0]["key"] == "PROJ-1"
        assert stories[0]["sprint"] == "Sprint 8"

    def test_windows_line_endings(self):
        raw = b"Issue key,Sprint,Summary,Sprint\r\nPROJ-1,Sprint 4,Story,Sprint 2\r\n"
        stories, _ = parse_jira_csv(raw)
        assert stories[0]["sprint"] == "Sprint 4"

    def test_single_sprint_column_still_works(self):
        data = csv_bytes(
            "Issue key,Sprint",
            "PROJ-1,Y26.SP07",
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Sprint 7"


class TestExcelSyncEdgeCases:

    def test_key_with_leading_trailing_spaces_in_excel(self):
        """Keys with whitespace in Excel are still matched and skipped."""
        xl = make_excel(["Sprint", "Key"], rows=[["Sprint 1", "  PROJ-1  "]])
        stories = [{"key": "PROJ-1", "sprint": "Sprint 1"}]
        _, added, skipped = sync_excel(xl, stories)
        assert skipped == ["PROJ-1"]
        assert added == []

    def test_all_stories_already_exist(self):
        xl = make_excel(["Sprint", "Key"], rows=[
            ["Sprint 1", "PROJ-1"],
            ["Sprint 1", "PROJ-2"],
        ])
        stories = [{"key": "PROJ-1", "sprint": "Sprint 1"},
                   {"key": "PROJ-2", "sprint": "Sprint 1"}]
        _, added, skipped = sync_excel(xl, stories)
        assert added == []
        assert skipped == ["PROJ-1", "PROJ-2"]

    def test_all_stories_new(self):
        xl = make_excel(["Sprint", "Key"])
        stories = [{"key": "PROJ-10", "sprint": "Sprint 1"},
                   {"key": "PROJ-11", "sprint": "Sprint 1"}]
        _, added, skipped = sync_excel(xl, stories)
        assert added == ["PROJ-10", "PROJ-11"]
        assert skipped == []

    def test_wrong_sheet_name_raises(self):
        """If the Excel has no 'Functional Testing' sheet, raise a clear error."""
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        buf = io.BytesIO()
        wb.save(buf)
        with pytest.raises(RuntimeError, match="Functional Testing"):
            sync_excel(buf.getvalue(), [{"key": "PROJ-1", "sprint": "Sprint 1"}])


# ── Trailing blank rows (the real-world table gap bug) ────────────────────────

class TestTrailingBlankRows:
    """Excel tables often have styled empty rows below the last data row.
    Stories must be written immediately after the last non-empty row."""

    def _rows(self, xl_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xl_bytes))
        return list(wb.active.iter_rows(values_only=True))

    def test_new_sprint_no_gap_after_trailing_blanks(self):
        """Black separator and story appear directly after last data row."""
        xl = make_excel_with_trailing_blanks(
            ["Sprint", "Key"],
            [["Sprint 4", "PROJ-1"]],
            extra_blank_rows=15,
        )
        result, added, _ = sync_excel(xl, [{"key": "PROJ-10", "sprint": "Sprint 5"}])
        assert added == ["PROJ-10"]
        rows = self._rows(result)
        # Row 0 = header, Row 1 = PROJ-1, Row 2 = black separator, Row 3 = PROJ-10
        assert rows[1][1] == "PROJ-1"
        assert rows[2][1] is None       # black separator has no key
        assert rows[3][1] == "PROJ-10"

    def test_old_sprint_no_gap_after_trailing_blanks(self):
        """Old-sprint insertion also respects trailing blank rows."""
        xl = make_excel_with_trailing_blanks(
            ["Sprint", "Key"],
            [["Sprint 4", "PROJ-1"]],
            extra_blank_rows=10,
        )
        result, added, _ = sync_excel(xl, [{"key": "PROJ-2", "sprint": "Sprint 4"}])
        assert added == ["PROJ-2"]
        rows = self._rows(result)
        # header, PROJ-1, PROJ-2 (inserted) — no gap
        assert rows[2][1] == "PROJ-2"

    def test_only_header_row_no_gap(self):
        """Sheet with only headers + trailing blanks still lands stories at row 2."""
        xl = make_excel_with_trailing_blanks(["Sprint", "Key"], [], extra_blank_rows=5)
        result, added, _ = sync_excel(xl, [{"key": "PROJ-1", "sprint": "Sprint 1"}])
        assert added == ["PROJ-1"]
        rows = self._rows(result)
        # Row 0 = header, Row 1 = black sep, Row 2 = PROJ-1
        assert rows[2][1] == "PROJ-1"

    def test_max_row_vs_last_data_row_diverge(self):
        """Verify that styled empty rows push ws.max_row beyond actual data."""
        xl = make_excel_with_trailing_blanks(
            ["Sprint", "Key"],
            [["Sprint 4", "PROJ-1"]],
            extra_blank_rows=20,
        )
        wb = openpyxl.load_workbook(io.BytesIO(xl))
        ws = wb["Functional Testing"]
        assert ws.max_row > 2, "styled blanks should push max_row above 2"
        from app import _last_data_row
        assert _last_data_row(ws) == 2


# ── Date column and test-user column writing ──────────────────────────────────

class TestDateAndUserColumns:

    def _cell(self, xl_bytes, row, col):
        wb = openpyxl.load_workbook(io.BytesIO(xl_bytes))
        v = wb["Functional Testing"].cell(row=row, column=col).value
        # openpyxl reads date cells back as datetime — normalise for comparison
        if isinstance(v, datetime):
            return v.date()
        return v

    def test_date_tested_header_writes_today(self):
        xl = make_excel(["Sprint", "Key", "Date Tested"])
        result, _, _ = sync_excel(xl, [{"key": "PROJ-1", "sprint": "Sprint 1"}])
        # new story is at row 3 (row 2 = black separator)
        assert self._cell(result, 3, 3) == date.today()

    def test_test_date_header_writes_today(self):
        xl = make_excel(["Sprint", "Key", "Test Date"])
        result, _, _ = sync_excel(xl, [{"key": "PROJ-1", "sprint": "Sprint 1"}])
        assert self._cell(result, 3, 3) == date.today()

    def test_date_header_case_insensitive(self):
        xl = make_excel(["Sprint", "Key", "DATE TESTED"])
        result, _, _ = sync_excel(xl, [{"key": "PROJ-1", "sprint": "Sprint 1"}])
        assert self._cell(result, 3, 3) == date.today()

    def test_today_date_object_header_writes_today(self):
        """A header cell that IS today's date as a date object is detected."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Functional Testing"
        ws.cell(row=1, column=1).value = "Sprint"
        ws.cell(row=1, column=2).value = "Key"
        ws.cell(row=1, column=3).value = date.today()
        buf = io.BytesIO()
        wb.save(buf)
        result, _, _ = sync_excel(buf.getvalue(), [{"key": "PROJ-1", "sprint": "Sprint 1"}])
        assert self._cell(result, 3, 3) == date.today()

    def test_yesterday_date_header_not_used_as_date_col(self):
        """A date object for yesterday in the header must NOT be used as today's col."""
        yesterday = date.today() - timedelta(days=1)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Functional Testing"
        ws.cell(row=1, column=1).value = "Sprint"
        ws.cell(row=1, column=2).value = "Key"
        ws.cell(row=1, column=3).value = yesterday  # old date column
        buf = io.BytesIO()
        wb.save(buf)
        result, _, _ = sync_excel(buf.getvalue(), [{"key": "PROJ-1", "sprint": "Sprint 1"}])
        # Column 3 should NOT have today's date written (it's yesterday's column)
        assert self._cell(result, 3, 3) is None

    def test_test_user_header_writes_peter(self):
        xl = make_excel(["Sprint", "Key", "Test User"])
        result, _, _ = sync_excel(xl, [{"key": "PROJ-1", "sprint": "Sprint 1"}])
        assert self._cell(result, 3, 3) == "Peter"

    def test_tester_header_writes_peter(self):
        xl = make_excel(["Sprint", "Key", "Tester"])
        result, _, _ = sync_excel(xl, [{"key": "PROJ-1", "sprint": "Sprint 1"}])
        assert self._cell(result, 3, 3) == "Peter"

    def test_peter_capitalised_in_old_sprint_insert(self):
        """Peter is capital even when inserted inline (not via append path)."""
        xl = make_excel(["Sprint", "Key", "Test User"],
                        rows=[["Sprint 4", "PROJ-1", None]])
        result, _, _ = sync_excel(xl, [{"key": "PROJ-2", "sprint": "Sprint 4"}])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Functional Testing"]
        keys = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
        peter_row = next(r for r in range(2, ws.max_row + 1)
                         if ws.cell(row=r, column=2).value == "PROJ-2")
        assert ws.cell(row=peter_row, column=3).value == "Peter"

    def test_no_date_or_user_col_still_works(self):
        """Sync must succeed even when the Excel has no date/user columns."""
        xl = make_excel(["Sprint", "Key"])
        result, added, _ = sync_excel(xl, [{"key": "PROJ-1", "sprint": "Sprint 1"}])
        assert added == ["PROJ-1"]


# ── Black separator fill ──────────────────────────────────────────────────────

class TestBlackSeparator:

    def test_separator_row_fill_is_black(self):
        """The separator row must have black fill on all cells."""
        xl = make_excel(["Sprint", "Key", "Info"],
                        rows=[["Sprint 4", "PROJ-1", "x"]])
        result, _, _ = sync_excel(xl, [{"key": "PROJ-10", "sprint": "Sprint 5"}])
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Functional Testing"]
        # separator is at row 3 (row 2 = PROJ-1, row 3 = sep, row 4 = PROJ-10)
        sep_row = 3
        for col in range(1, 4):
            fill = ws.cell(row=sep_row, column=col).fill
            assert fill.fgColor.rgb.upper().endswith("000000"), \
                f"col {col} fill should be black, got {fill.fgColor.rgb}"

    def test_no_separator_when_all_old_sprint(self):
        """No black separator row added when all stories match existing sprints."""
        xl = make_excel(["Sprint", "Key"], rows=[["Sprint 4", "PROJ-1"]])
        result, added, _ = sync_excel(xl, [{"key": "PROJ-2", "sprint": "Sprint 4"}])
        assert added == ["PROJ-2"]
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Functional Testing"]
        # header, PROJ-1, PROJ-2 — exactly 3 rows, no black sep
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3

    def test_only_one_separator_for_multiple_new_sprints(self):
        """Multiple new sprints share a single black separator row."""
        xl = make_excel(["Sprint", "Key"], rows=[["Sprint 3", "PROJ-1"]])
        stories = [
            {"key": "PROJ-10", "sprint": "Sprint 5"},
            {"key": "PROJ-11", "sprint": "Sprint 6"},
        ]
        result, added, _ = sync_excel(xl, stories)
        assert set(added) == {"PROJ-10", "PROJ-11"}
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Functional Testing"]
        rows = list(ws.iter_rows(values_only=True))
        # header, PROJ-1, black_sep, PROJ-10, PROJ-11 → exactly 5 rows
        assert len(rows) == 5
        assert rows[2][1] is None   # black sep
        assert rows[3][1] == "PROJ-10"
        assert rows[4][1] == "PROJ-11"


# ── Multiple sprint insertion ordering ───────────────────────────────────────

class TestMultiSprintInsertion:

    def _rows(self, xl_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(xl_bytes))
        return list(wb.active.iter_rows(values_only=True))

    def test_two_old_sprints_inserted_in_correct_blocks(self):
        """Stories for Sprint 3 and Sprint 4 are inserted near their own blocks."""
        xl = make_excel(["Sprint", "Key"], rows=[
            ["Sprint 3", "PROJ-1"],
            ["Sprint 3", "PROJ-2"],
            ["Sprint 4", "PROJ-3"],
            ["Sprint 4", "PROJ-4"],
        ])
        stories = [
            {"key": "PROJ-5", "sprint": "Sprint 3"},
            {"key": "PROJ-6", "sprint": "Sprint 4"},
        ]
        result, added, _ = sync_excel(xl, stories)
        assert set(added) == {"PROJ-5", "PROJ-6"}
        rows = self._rows(result)
        keys = [r[1] for r in rows[1:]]
        # Sprint 3 block: PROJ-1, PROJ-2, PROJ-5 — then Sprint 4 block: PROJ-3, PROJ-4, PROJ-6
        assert keys.index("PROJ-5") < keys.index("PROJ-3"), \
            "PROJ-5 (Sprint 3) must appear before PROJ-3 (Sprint 4)"
        assert keys.index("PROJ-6") > keys.index("PROJ-4"), \
            "PROJ-6 (Sprint 4) must appear after PROJ-4"
        assert keys.index("PROJ-5") == keys.index("PROJ-2") + 1, \
            "PROJ-5 must immediately follow PROJ-2"

    def test_insertion_order_within_sprint_preserved(self):
        """Multiple stories added to the same sprint maintain input order."""
        xl = make_excel(["Sprint", "Key"], rows=[["Sprint 4", "PROJ-1"]])
        stories = [
            {"key": "PROJ-2", "sprint": "Sprint 4"},
            {"key": "PROJ-3", "sprint": "Sprint 4"},
            {"key": "PROJ-4", "sprint": "Sprint 4"},
        ]
        result, added, _ = sync_excel(xl, stories)
        assert added == ["PROJ-2", "PROJ-3", "PROJ-4"]
        rows = self._rows(result)
        keys = [r[1] for r in rows[1:]]
        assert keys == ["PROJ-1", "PROJ-2", "PROJ-3", "PROJ-4"]

    def test_mixed_sprints_correct_final_layout(self):
        """Combined old-sprint insertion + new-sprint append produces correct layout."""
        xl = make_excel(["Sprint", "Key"], rows=[
            ["Sprint 3", "PROJ-A"],
            ["Sprint 4", "PROJ-B"],
        ])
        stories = [
            {"key": "PROJ-C", "sprint": "Sprint 3"},   # old
            {"key": "PROJ-D", "sprint": "Sprint 4"},   # old
            {"key": "PROJ-E", "sprint": "Sprint 5"},   # new
        ]
        result, added, _ = sync_excel(xl, stories)
        assert set(added) == {"PROJ-C", "PROJ-D", "PROJ-E"}
        rows = self._rows(result)
        keys = [r[1] for r in rows]
        assert keys.index("PROJ-C") < keys.index("PROJ-B"), "PROJ-C in Sprint 3 block"
        assert keys.index("PROJ-D") > keys.index("PROJ-B"), "PROJ-D after Sprint 4 start"
        assert keys.index("PROJ-E") > keys.index("PROJ-D"), "new sprint after old sprints"
        assert None in keys, "black separator row present"


# ── Sprint normalisation edge cases ──────────────────────────────────────────

class TestSprintNormalisationEdgeCases:

    def _sprint(self, raw):
        data = f"Issue key,Sprint\nPROJ-1,{raw}\n".encode()
        stories, _ = parse_jira_csv(data)
        return stories[0]["sprint"]

    def test_sprint_with_leading_zeros_sp001(self):
        assert self._sprint("SP001") == "Sprint 1"

    def test_sprint_100(self):
        assert self._sprint("Sprint 100") == "Sprint 100"

    def test_sprint_sp10(self):
        assert self._sprint("SP10") == "Sprint 10"

    def test_sprint_all_caps(self):
        assert self._sprint("SPRINT 7") == "Sprint 7"

    def test_sprint_mixed_case_yy_sp(self):
        assert self._sprint("Y26.SP12") == "Sprint 12"

    def test_sprint_with_surrounding_spaces(self):
        data = b"Issue key,Sprint\nPROJ-1,  Sprint 3  \n"
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Sprint 3"

    def test_no_sprint_match_returns_raw(self):
        """If the sprint value has no recognisable pattern, return it as-is."""
        data = b"Issue key,Sprint\nPROJ-1,Backlog\n"
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Backlog"

    def test_empty_sprint_value(self):
        data = b"Issue key,Sprint\nPROJ-1,\n"
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == ""

    def test_story_with_empty_sprint_gets_black_separator(self):
        """Story with sprint='' is treated as new sprint (won't match any existing)."""
        xl = make_excel(["Sprint", "Key"], rows=[["Sprint 4", "PROJ-1"]])
        result, added, _ = sync_excel(xl, [{"key": "PROJ-2", "sprint": ""}])
        assert added == ["PROJ-2"]
        rows = list(openpyxl.load_workbook(io.BytesIO(result)).active.iter_rows(values_only=True))
        # black separator + PROJ-2 after PROJ-1
        assert len(rows) == 4


# ── CSV parser robustness ─────────────────────────────────────────────────────

class TestCsvParserRobustness:

    def test_all_rows_invalid_returns_empty(self):
        data = csv_bytes(
            "Issue key,Sprint",
            "Product Owner,",
            "Epic,,",
            ",Sprint 1",
        )
        stories, _ = parse_jira_csv(data)
        assert stories == []

    def test_single_valid_row(self):
        data = csv_bytes("Issue key,Sprint", "PROJ-1,Sprint 1")
        stories, _ = parse_jira_csv(data)
        assert len(stories) == 1

    def test_very_large_sprint_number(self):
        data = csv_bytes("Issue key,Sprint", "PROJ-1,Sprint 999")
        stories, _ = parse_jira_csv(data)
        assert stories[0]["sprint"] == "Sprint 999"

    def test_key_with_numbers_in_project_code(self):
        data = csv_bytes("Issue key,Sprint", "PROJ2024-1,Sprint 1")
        stories, _ = parse_jira_csv(data)
        assert stories[0]["key"] == "PROJ2024-1"

    def test_key_lowercase_accepted(self):
        data = csv_bytes("Issue key,Sprint", "proj-1,Sprint 1")
        stories, _ = parse_jira_csv(data)
        assert stories[0]["key"] == "proj-1"

    def test_whitespace_only_rows_skipped(self):
        data = b"Issue key,Sprint\nPROJ-1,Sprint 1\n   \n\t\nPROJ-2,Sprint 2\n"
        stories, _ = parse_jira_csv(data)
        assert len(stories) == 2

    def test_header_row_only_returns_empty(self):
        data = b"Issue key,Sprint\n"
        with pytest.raises(RuntimeError):
            parse_jira_csv(data)

    def test_summary_column_present_ignored(self):
        data = csv_bytes(
            "Issue key,Summary,Sprint",
            "PROJ-1,A very long summary with commas? no,Sprint 2",
        )
        stories, _ = parse_jira_csv(data)
        assert stories[0]["key"] == "PROJ-1"
        assert stories[0]["sprint"] == "Sprint 2"


# ── Route edge cases ──────────────────────────────────────────────────────────

class TestRouteEdgeCases:

    @pytest.fixture
    def client(self):
        app.config["TESTING"] = True
        with app.test_client() as c:
            yield c

    def test_all_stories_returned_in_fetch_response(self, client):
        xl = make_excel(["Sprint", "Key"])
        stories = [
            {"key": "PROJ-1", "sprint": "Sprint 1"},
            {"key": "PROJ-2", "sprint": "Sprint 1"},
        ]
        resp = client.post(
            "/sync",
            data={
                "stories_json": (io.BytesIO(json.dumps(stories).encode()), "s.json"),
                "excel_file": (io.BytesIO(xl), "tracker.xlsx"),
            },
            content_type="multipart/form-data",
            headers={"X-Requested-With": "fetch"},
        )
        data = json.loads(resp.data)
        assert set(data["all_stories"]) == {"PROJ-1", "PROJ-2"}

    def test_pasted_json_input_accepted(self, client):
        xl = make_excel(["Sprint", "Key"])
        stories = [{"key": "PROJ-99", "sprint": "Sprint 9"}]
        resp = client.post(
            "/sync",
            data={
                "stories_json_text": json.dumps(stories),
                "excel_file": (io.BytesIO(xl), "tracker.xlsx"),
            },
            content_type="multipart/form-data",
            headers={"X-Requested-With": "fetch"},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert "PROJ-99" in data["added"]

    def test_csv_uploaded_directly_to_sync(self, client):
        xl = make_excel(["Sprint", "Key"])
        csv_data = csv_bytes("Issue key,Sprint", "PROJ-5,Sprint 1")
        resp = client.post(
            "/sync",
            data={
                "jira_csv": (io.BytesIO(csv_data), "export.csv"),
                "excel_file": (io.BytesIO(xl), "tracker.xlsx"),
            },
            content_type="multipart/form-data",
            headers={"X-Requested-With": "fetch"},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert "PROJ-5" in data["added"]

    def test_invalid_pasted_json_redirects(self, client):
        xl = make_excel(["Sprint", "Key"])
        resp = client.post(
            "/sync",
            data={
                "stories_json_text": "{this is not json}",
                "excel_file": (io.BytesIO(xl), "tracker.xlsx"),
            },
            content_type="multipart/form-data",
        )
        assert resp.status_code == 303
        assert "error" in resp.headers["Location"]

    def test_invalid_json_file_redirects(self, client):
        xl = make_excel(["Sprint", "Key"])
        resp = client.post(
            "/sync",
            data={
                "stories_json": (io.BytesIO(b"not json at all"), "s.json"),
                "excel_file": (io.BytesIO(xl), "tracker.xlsx"),
            },
            content_type="multipart/form-data",
        )
        assert resp.status_code == 303

    def test_empty_json_array_redirects(self, client):
        xl = make_excel(["Sprint", "Key"])
        resp = client.post(
            "/sync",
            data={
                "stories_json": (io.BytesIO(b"[]"), "s.json"),
                "excel_file": (io.BytesIO(xl), "tracker.xlsx"),
            },
            content_type="multipart/form-data",
        )
        assert resp.status_code == 303

    def test_all_stories_skipped_returns_200_with_empty_added(self, client):
        xl = make_excel(["Sprint", "Key"], rows=[["Sprint 1", "PROJ-1"]])
        stories = [{"key": "PROJ-1", "sprint": "Sprint 1"}]
        resp = client.post(
            "/sync",
            data={
                "stories_json": (io.BytesIO(json.dumps(stories).encode()), "s.json"),
                "excel_file": (io.BytesIO(xl), "tracker.xlsx"),
            },
            content_type="multipart/form-data",
            headers={"X-Requested-With": "fetch"},
        )
        assert resp.status_code == 200
        data = json.loads(resp.data)
        assert data["added"] == []
        assert data["skipped"] == ["PROJ-1"]

    def test_filename_contains_today(self, client):
        xl = make_excel(["Sprint", "Key"])
        stories = [{"key": "PROJ-1", "sprint": "Sprint 1"}]
        resp = client.post(
            "/sync",
            data={
                "stories_json": (io.BytesIO(json.dumps(stories).encode()), "s.json"),
                "excel_file": (io.BytesIO(xl), "tracker.xlsx"),
            },
            content_type="multipart/form-data",
            headers={"X-Requested-With": "fetch"},
        )
        data = json.loads(resp.data)
        today_str = date.today().strftime("%Y%m%d")
        assert today_str in data["filename"]

    def test_excel_b64_decodes_to_valid_xlsx(self, client):
        import base64
        xl = make_excel(["Sprint", "Key"])
        stories = [{"key": "PROJ-1", "sprint": "Sprint 1"}]
        resp = client.post(
            "/sync",
            data={
                "stories_json": (io.BytesIO(json.dumps(stories).encode()), "s.json"),
                "excel_file": (io.BytesIO(xl), "tracker.xlsx"),
            },
            content_type="multipart/form-data",
            headers={"X-Requested-With": "fetch"},
        )
        data = json.loads(resp.data)
        raw = base64.b64decode(data["excel_b64"])
        # Should be a valid xlsx (PK zip magic bytes)
        assert raw[:2] == b"PK"
